"""
Scan Unturned bundles and export item IDs, names, descriptions, and GUIDs.

Scanner automatically discovers folders that contain item, effect, tree, or vehicle
subfolders, skips ignored paths such as Mythics, and sorts output into category
worksheets.

Author: Tony Seed
2026
"""

from __future__ import annotations

import argparse
import re
from collections import deque
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from pathlib import Path


DEFAULT_BUNDLES_FOLDER = Path(r"D:\SteamLibrary\steamapps\common\Unturned\Bundles")
DEFAULT_WORKSHOP_FOLDER = Path(r"D:\SteamLibrary\steamapps\workshop\content\304930")
DEFAULT_OUTPUT_FILE = Path(r"C:\Users\Tony\Documents\UnturnedIDList\UnturnedItems.xlsx")
DEFAULT_IGNORED_FOLDERS = {"Mythics"}

ITEM_ID_RE = re.compile(r"^ID\s+(\d+)\s*(?://.*)?$")
GUID_RE = re.compile(r"^GUID\s+(.+?)\s*(?://.*)?$")
NAME_RE = re.compile(r"^Name\s+(.+)$")
DESCRIPTION_RE = re.compile(r"^Description\s+(.+)$")
HEADER_ROW = ["(Leave Empty)", "ID", "Name", "Description", "(Leave Empty)", "GUID"]
ALL_ITEMS_HEADER_ROW = [
    "Category",
    "Folder",
    "ID",
    "Name",
    "Description",
    "GUID",
]
INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?:\[\]]")
MAX_EXCEL_SHEET_NAME_LENGTH = 31


@dataclass(frozen=True)
class ItemRecord:
    """A single scraped Unturned item row."""

    category: str
    folder_name: str
    item_id: str
    name: str
    description: str
    guid: str


def normalize_names(names: Iterable[str]) -> set[str]:
    """Return case-insensitive folder names with empty values removed."""
    return {name.strip().casefold() for name in names if name and name.strip()}


def should_ignore(path: Path, ignored_folders: set[str]) -> bool:
    """Return True when any part of a path matches an ignored folder name."""
    return any(part.casefold() in ignored_folders for part in path.parts)


def find_asset_dat(path: Path) -> Path | None:
    """Return the best .dat metadata file for an asset folder.

    Items usually use ``FolderName/FolderName.dat`` plus ``English.dat``, while
    Vehicles, Trees, and Effects can use different .dat names or skip
    ``English.dat`` entirely. Prefer the matching folder-name .dat when it
    exists, otherwise use the first non-English .dat file in the folder.
    """
    matching_dat = path / f"{path.name}.dat"
    if matching_dat.is_file():
        return matching_dat

    try:
        dat_files = sorted(
            file
            for file in path.glob("*.dat")
            if file.name.casefold() not in {"english.dat", "masterbundle.dat"}
        )
    except OSError as error:
        print(f"Error scanning {path}: {error}")
        return None

    return dat_files[0] if dat_files else None


def is_item_folder(path: Path) -> bool:
    """Return True when a folder contains a scrapeable asset .dat file."""
    return find_asset_dat(path) is not None


def contains_item_folders(path: Path, ignored_folders: set[str]) -> bool:
    """Return True when a folder has at least one direct child asset folder."""
    try:
        children = path.iterdir()
    except OSError as error:
        print(f"Error scanning {path}: {error}")
        return False

    return any(
        child.is_dir() and not should_ignore(child, ignored_folders) and is_item_folder(child)
        for child in children
    )


def discover_export_folders(root_folders: Sequence[Path], ignored_folders: set[str]) -> list[Path]:
    """Find asset category folders that should become worksheets in the export workbook."""
    export_folders: list[Path] = []
    pending = deque()

    # Add all root folders to the pending queue
    for folder in root_folders:
        if folder.exists():
            pending.append(folder)
        else:
            print(f"Skipping missing folder: {folder}")

    while pending:
        current = pending.popleft()
        if should_ignore(current, ignored_folders):
            continue

        if contains_item_folders(current, ignored_folders):
            export_folders.append(current)
            continue

        try:
            child_folders = sorted(child for child in current.iterdir() if child.is_dir())
        except OSError as error:
            print(f"Error scanning {current}: {error}")
            continue

        pending.extend(child for child in child_folders if not should_ignore(child, ignored_folders))

    return export_folders


def category_name(root_folder: Path, bundles_folder: Path) -> str:
    """Return a readable category name from a discovered folder path."""
    try:
        return " / ".join(root_folder.relative_to(bundles_folder).parts)
    except ValueError:
        return root_folder.name


def worksheet_title(title: str, used_titles: set[str]) -> str:
    """Create a safe, unique Excel worksheet title."""
    title = INVALID_SHEET_CHARS_RE.sub("-", title).strip() or "Items"
    title = title[:MAX_EXCEL_SHEET_NAME_LENGTH]
    candidate = title
    suffix = 2

    while candidate in used_titles:
        suffix_text = f" ({suffix})"
        candidate = f"{title[:MAX_EXCEL_SHEET_NAME_LENGTH - len(suffix_text)]}{suffix_text}"
        suffix += 1

    used_titles.add(candidate)
    return candidate


def read_item_metadata(item_dat: Path) -> tuple[str, str]:
    """Return the top-level numeric item ID and GUID from an item .dat file."""
    item_id = ""
    guid = ""
    block_depth = 0

    with item_dat.open("r", encoding="utf-8-sig", errors="ignore") as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line:
                continue

            # Only read keys from the root of the .dat file. Recipe/crafting
            # sections use [ ] and { } blocks and can contain their own `ID`
            # keys that point at GUIDs for required input items.
            if block_depth == 0:
                if not item_id:
                    id_match = ITEM_ID_RE.match(line)
                    if id_match:
                        item_id = id_match.group(1)

                if not guid:
                    guid_match = GUID_RE.match(line)
                    if guid_match:
                        guid = guid_match.group(1).strip()

                if item_id and guid:
                    return item_id, guid

            block_depth += line.count("[") + line.count("{")
            block_depth -= line.count("]") + line.count("}")
            block_depth = max(block_depth, 0)

    return item_id, guid


def read_english_metadata(english_dat: Path) -> tuple[str, str]:
    """Return item name and description from an English.dat file."""
    name = ""
    description = ""

    with english_dat.open("r", encoding="utf-8-sig", errors="ignore") as file:
        for raw_line in file:
            line = raw_line.strip()

            if not name:
                name_match = NAME_RE.match(line)
                if name_match:
                    name = name_match.group(1).strip()
                    continue

            if not description:
                description_match = DESCRIPTION_RE.match(line)
                if description_match:
                    description = description_match.group(1).strip()

    return name, description


def read_asset_record(folder_path: Path, category: str) -> ItemRecord | None:
    """Scrape one asset folder into an ItemRecord."""
    item_dat = find_asset_dat(folder_path)
    if item_dat is None:
        return None

    english_dat = folder_path / "English.dat"

    try:
        item_id, guid = read_item_metadata(item_dat)
    except OSError as error:
        print(f"Error reading {item_dat}: {error}")
        return None

    if english_dat.exists():
        try:
            name, description = read_english_metadata(english_dat)
        except OSError as error:
            print(f"Error reading {english_dat}: {error}")
            name = ""
            description = ""
    else:
        name = folder_path.name.replace("_", " ")
        description = ""

    return ItemRecord(
        category=category,
        folder_name=folder_path.name,
        item_id=item_id,
        name=name,
        description=description,
        guid=guid,
    )


def collect_item_records(
    root_folder: Path,
    bundles_folder: Path,
    ignored_folders: set[str],
) -> list[ItemRecord]:
    """Scrape all assets from one category folder."""
    records: list[ItemRecord] = []
    category = category_name(root_folder, bundles_folder)

    try:
        item_folders = sorted(child for child in root_folder.iterdir() if child.is_dir())
    except OSError as error:
        print(f"Error scanning {root_folder}: {error}")
        return records

    for folder_path in item_folders:
        if should_ignore(folder_path, ignored_folders) or not is_item_folder(folder_path):
            continue

        record = read_asset_record(folder_path, category)
        if record is not None:
            records.append(record)

    return sorted(records, key=lambda record: (record.name.casefold(), record.item_id))


def collect_selected_asset_records(
    asset_folders: Sequence[Path],
    bundles_folders: Sequence[Path],
    ignored_folders: set[str],
) -> dict[str, list[ItemRecord]]:
    """Scrape checked asset folders and group them by category."""
    records_by_category: dict[str, list[ItemRecord]] = {}

    for folder_path in sorted(asset_folders):
        if should_ignore(folder_path, ignored_folders) or not is_item_folder(folder_path):
            continue

        parent_folder = folder_path.parent
        # Find the right bundles folder to use for category name
        category = None
        for bundles_folder in bundles_folders:
            try:
                parent_folder.relative_to(bundles_folder)
                category = category_name(parent_folder, bundles_folder)
                break
            except ValueError:
                continue
        
        if category is None:
            category = parent_folder.name
        
        record = read_asset_record(folder_path, category)
        if record is None:
            continue

        records_by_category.setdefault(category, []).append(record)

    for records in records_by_category.values():
        records.sort(key=lambda record: (record.name.casefold(), record.item_id))

    return records_by_category


def append_category_sheet(sheet, records: Sequence[ItemRecord]) -> None:
    """Append one category worksheet using the requested blank/ID/name layout."""
    sheet.append(HEADER_ROW)
    for record in records:
        sheet.append(["", record.item_id, record.name, record.description, "", record.guid])


def append_all_items_sheet(sheet, records: Sequence[ItemRecord]) -> None:
    """Append an automatically sorted all-items worksheet."""
    sheet.append(ALL_ITEMS_HEADER_ROW)
    for record in sorted(records, key=lambda item: (item.category, item.name.casefold(), item.item_id)):
        sheet.append(
            [record.category, record.folder_name, record.item_id, record.name, record.description, record.guid]
        )


def export_workbook(
    root_folders: Sequence[Path],
    output_file: Path,
    export_folders: Sequence[Path],
    ignored_folders: set[str],
) -> int:
    """Export selected category folders to an XLSX workbook and return row count."""
    from openpyxl import Workbook

    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "All Items"
    used_titles = {summary_sheet.title}
    records_by_category: dict[str, list[ItemRecord]] = {}
    all_records: list[ItemRecord] = []

    # Use the first root folder as the reference for category names
    primary_folder = root_folders[0] if root_folders else None

    for root_folder in export_folders:
        # Find which root folder this export folder belongs to
        category_ref = primary_folder
        for root in root_folders:
            try:
                root_folder.relative_to(root)
                category_ref = root
                break
            except ValueError:
                continue

        if category_ref:
            records = collect_item_records(root_folder, category_ref, ignored_folders)
            category = category_name(root_folder, category_ref)
        else:
            records = []
            category = root_folder.name

        records_by_category[category] = records
        all_records.extend(records)
        print(f"Exported {len(records)} rows from {root_folder}")

    append_all_items_sheet(summary_sheet, all_records)

    for category, records in sorted(records_by_category.items()):
        sheet = workbook.create_sheet(title=worksheet_title(category, used_titles))
        append_category_sheet(sheet, records)

    if not export_folders:
        empty_sheet = workbook.create_sheet(title="No Items Found")
        append_category_sheet(empty_sheet, [])

    workbook.save(output_file)
    return len(all_records)


def export_selected_asset_workbook(
    root_folders: Sequence[Path],
    output_file: Path,
    asset_folders: Sequence[Path],
    ignored_folders: set[str],
) -> int:
    """Export individually checked asset folders to an XLSX workbook."""
    from openpyxl import Workbook

    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "All Items"
    used_titles = {summary_sheet.title}

    records_by_category = collect_selected_asset_records(asset_folders, root_folders, ignored_folders)
    all_records = [record for records in records_by_category.values() for record in records]
    append_all_items_sheet(summary_sheet, all_records)

    for category, records in sorted(records_by_category.items()):
        sheet = workbook.create_sheet(title=worksheet_title(category, used_titles))
        append_category_sheet(sheet, records)

    workbook.save(output_file)
    return len(all_records)


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    """Parse command-line options for scanning, exporting, and launching the app."""
    parser = argparse.ArgumentParser(
        description="Scan Unturned bundle folders and export item metadata to XLSX."
    )
    parser.add_argument(
        "--bundles-folder",
        type=Path,
        default=DEFAULT_BUNDLES_FOLDER,
        help="Path to the Unturned Bundles folder.",
    )
    parser.add_argument(
        "--workshop-folder",
        type=Path,
        default=DEFAULT_WORKSHOP_FOLDER,
        help="Path to the Steam Workshop content folder for Unturned.",
    )
    parser.add_argument(
        "--output-file",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help="Path to the XLSX file to create.",
    )
    parser.add_argument(
        "--ignore",
        action="append",
        default=[],
        help="Folder name to ignore anywhere in the scan. Can be used multiple times.",
    )
    parser.add_argument(
        "--include-default-ignores",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Include built-in ignored folders like Mythics.",
    )
    parser.add_argument(
        "--scan-only",
        action="store_true",
        help="Only print the worksheet paths that would be exported; do not create XLSX.",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Launch a checkbox-based desktop app for scanning and exporting.",
    )
    return parser.parse_args(argv)


def ignored_folders_from_args(args: argparse.Namespace) -> set[str]:
    """Build the effective ignored-folder set from CLI options."""
    ignored_folders = normalize_names(args.ignore)
    if args.include_default_ignores:
        ignored_folders.update(normalize_names(DEFAULT_IGNORED_FOLDERS))
    return ignored_folders


def launch_gui(default_bundles_folder: Path, default_workshop_folder: Path, default_output_file: Path) -> None:
    """Launch a small Tkinter app for tree-based scanning and exporting."""
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    CHECKED = "☑"
    UNCHECKED = "☐"

    class ExportApp:
        def __init__(self, root: tk.Tk) -> None:
            self.root = root
            self.root.title("Unturned ID Exporter")
            self.bundles_folder = tk.StringVar(value=str(default_bundles_folder))
            self.workshop_folder = tk.StringVar(value=str(default_workshop_folder))
            self.output_file = tk.StringVar(value=str(default_output_file))
            self.ignore_text = tk.StringVar(value=", ".join(sorted(DEFAULT_IGNORED_FOLDERS)))
            self.status = tk.StringVar(value="Choose folders, then click Scan folders.")
            self.node_paths: dict[str, Path] = {}
            self.node_labels: dict[str, str] = {}
            self.asset_nodes: set[str] = set()
            self.checked_nodes: set[str] = set()

            self.frame = ttk.Frame(root, padding=12)
            self.frame.grid(row=0, column=0, sticky="nsew")
            root.columnconfigure(0, weight=1)
            root.rowconfigure(0, weight=1)
            self.frame.columnconfigure(1, weight=1)
            self.frame.rowconfigure(5, weight=1)

            ttk.Label(self.frame, text="Bundles folder").grid(row=0, column=0, sticky="w")
            ttk.Entry(self.frame, textvariable=self.bundles_folder).grid(row=0, column=1, sticky="ew")
            ttk.Button(self.frame, text="Browse", command=self.choose_bundles_folder).grid(row=0, column=2)

            ttk.Label(self.frame, text="Workshop folder").grid(row=1, column=0, sticky="w")
            ttk.Entry(self.frame, textvariable=self.workshop_folder).grid(row=1, column=1, sticky="ew")
            ttk.Button(self.frame, text="Browse", command=self.choose_workshop_folder).grid(row=1, column=2)

            ttk.Label(self.frame, text="Output XLSX").grid(row=2, column=0, sticky="w")
            ttk.Entry(self.frame, textvariable=self.output_file).grid(row=2, column=1, sticky="ew")
            ttk.Button(self.frame, text="Save as", command=self.choose_output_file).grid(row=2, column=2)

            ttk.Label(self.frame, text="Ignore folders").grid(row=3, column=0, sticky="w")
            ttk.Entry(self.frame, textvariable=self.ignore_text).grid(row=3, column=1, sticky="ew")
            ttk.Label(self.frame, text="Comma-separated").grid(row=3, column=2, sticky="w")

            self.button_frame = ttk.Frame(self.frame)
            self.button_frame.grid(row=4, column=0, columnspan=3, sticky="w")
            ttk.Button(self.button_frame, text="Scan folders", command=self.scan_folders).grid(row=0, column=0)
            ttk.Button(self.button_frame, text="Export checked", command=self.export_checked).grid(row=0, column=1)
            ttk.Button(self.button_frame, text="Check all", command=self.check_all).grid(row=0, column=2)
            ttk.Button(self.button_frame, text="Uncheck all", command=self.uncheck_all).grid(row=0, column=3)

            self.tree = ttk.Treeview(self.frame, show="tree", selectmode="browse")
            self.scrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=self.tree.yview)
            self.tree.configure(yscrollcommand=self.scrollbar.set)
            self.tree.grid(row=5, column=0, columnspan=2, sticky="nsew")
            self.scrollbar.grid(row=5, column=2, sticky="ns")
            self.tree.bind("<ButtonRelease-1>", self.toggle_clicked_node)
            self.tree.bind("<space>", self.toggle_selected_node)

            ttk.Label(self.frame, textvariable=self.status).grid(row=6, column=0, columnspan=3, sticky="w")

        def ignored_folders(self) -> set[str]:
            return normalize_names(self.ignore_text.get().split(","))

        def choose_bundles_folder(self) -> None:
            folder = filedialog.askdirectory(title="Select Bundles folder")
            if folder:
                self.bundles_folder.set(folder)

        def choose_workshop_folder(self) -> None:
            folder = filedialog.askdirectory(title="Select Workshop folder")
            if folder:
                self.workshop_folder.set(folder)

        def choose_output_file(self) -> None:
            file = filedialog.asksaveasfilename(
                title="Save XLSX as",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
            if file:
                self.output_file.set(file)

        def scan_folders(self) -> None:
            self.tree.delete(*self.tree.get_children())
            self.node_paths.clear()
            self.node_labels.clear()
            self.asset_nodes.clear()
            self.checked_nodes.clear()

            root_folders = [Path(self.bundles_folder.get()), Path(self.workshop_folder.get())]
            export_folders = discover_export_folders(root_folders, self.ignored_folders())

            for category_folder in export_folders:
                # Find which root folder this belongs to
                category_ref = root_folders[0]
                for root in root_folders:
                    try:
                        category_folder.relative_to(root)
                        category_ref = root
                        break
                    except ValueError:
                        continue

                category_node = self.insert_path_nodes(category_folder, category_ref)
                for asset_folder in self.asset_folders_for_category(category_folder):
                    asset_node = self.tree.insert(
                        category_node,
                        "end",
                        text=f"☑ {asset_folder.name}",
                        open=False,
                    )
                    self.node_paths[asset_node] = asset_folder
                    self.node_labels[asset_node] = asset_folder.name
                    self.asset_nodes.add(asset_node)
                    self.checked_nodes.add(asset_node)

            for node in self.tree.get_children():
                self.refresh_parent_checks(node)
                self.tree.item(node, open=True)

            self.status.set(f"Found {len(self.asset_nodes)} assets. Check or uncheck any folder level.")

        def insert_path_nodes(self, folder: Path, bundles_folder: Path) -> str:
            parent = ""
            current_path = bundles_folder
            for part in folder.relative_to(bundles_folder).parts:
                current_path = current_path / part
                existing = self.find_child(parent, current_path)
                if existing:
                    parent = existing
                    continue

                node = self.tree.insert(parent, "end", text=f"☑ {part}", open=True)
                self.node_paths[node] = current_path
                self.node_labels[node] = part
                self.checked_nodes.add(node)
                parent = node

            return parent

        def find_child(self, parent: str, path: Path) -> str:
            for child in self.tree.get_children(parent):
                if self.node_paths.get(child) == path:
                    return child
            return ""

        def asset_folders_for_category(self, category_folder: Path) -> list[Path]:
            try:
                children = sorted(child for child in category_folder.iterdir() if child.is_dir())
            except OSError as error:
                print(f"Error scanning {category_folder}: {error}")
                return []

            ignored = self.ignored_folders()
            return [child for child in children if not should_ignore(child, ignored) and is_item_folder(child)]

        def toggle_clicked_node(self, event: tk.Event) -> None:
            node = self.tree.identify_row(event.y)
            if node:
                self.toggle_node(node)

        def toggle_selected_node(self, event: tk.Event) -> str:
            selected = self.tree.selection()
            if selected:
                self.toggle_node(selected[0])
            return "break"

        def toggle_node(self, node: str) -> None:
            self.set_checked(node, node not in self.checked_nodes)
            self.refresh_ancestors(node)
            self.status.set(f"{len(self.selected_asset_folders())} assets checked for export.")

        def set_checked(self, node: str, checked: bool) -> None:
            if checked:
                self.checked_nodes.add(node)
            else:
                self.checked_nodes.discard(node)
            self.update_node_text(node)

            for child in self.tree.get_children(node):
                self.set_checked(child, checked)

        def refresh_ancestors(self, node: str) -> None:
            parent = self.tree.parent(node)
            while parent:
                self.refresh_parent_checks(parent)
                parent = self.tree.parent(parent)

        def refresh_parent_checks(self, node: str) -> bool:
            children = self.tree.get_children(node)
            if children:
                checked = all(self.refresh_parent_checks(child) for child in children)
                if checked:
                    self.checked_nodes.add(node)
                else:
                    self.checked_nodes.discard(node)
                self.update_node_text(node)
                return checked

            self.update_node_text(node)
            return node in self.checked_nodes

        def update_node_text(self, node: str) -> None:
            icon = CHECKED if node in self.checked_nodes else UNCHECKED
            self.tree.item(node, text=f"{icon} {self.node_labels[node]}")

        def selected_asset_folders(self) -> list[Path]:
            return [self.node_paths[node] for node in self.asset_nodes if node in self.checked_nodes]

        def check_all(self) -> None:
            for node in self.tree.get_children():
                self.set_checked(node, True)
            self.status.set(f"{len(self.selected_asset_folders())} assets checked for export.")

        def uncheck_all(self) -> None:
            for node in self.tree.get_children():
                self.set_checked(node, False)
            self.status.set("0 assets checked for export.")

        def export_checked(self) -> None:
            selected = self.selected_asset_folders()
            if not selected:
                messagebox.showwarning("Nothing selected", "Scan first, then keep at least one asset checked.")
                return

            root_folders = [Path(self.bundles_folder.get()), Path(self.workshop_folder.get())]
            row_count = export_selected_asset_workbook(
                root_folders,
                Path(self.output_file.get()),
                selected,
                self.ignored_folders(),
            )
            self.status.set(f"Exported {row_count} assets to {self.output_file.get()}")
            messagebox.showinfo("Done", f"Exported {row_count} assets.")

    root = tk.Tk()
    ExportApp(root)
    root.mainloop()


def main(argv: Sequence[str] | None = None) -> None:
    args = parse_args(argv)

    if args.gui:
        launch_gui(args.bundles_folder, args.workshop_folder, args.output_file)
        return

    root_folders = [args.bundles_folder, args.workshop_folder]
    ignored_folders = ignored_folders_from_args(args)
    export_folders = discover_export_folders(root_folders, ignored_folders)

    if args.scan_only:
        print("Folders that would be exported:")
        for folder in export_folders:
            print(folder)
        print(f"Total: {len(export_folders)}")
        return

    total_rows = export_workbook(root_folders, args.output_file, export_folders, ignored_folders)

    print("Done!")
    print(f"Exported {total_rows} total items")
    print(args.output_file)


if __name__ == "__main__":
    main()
